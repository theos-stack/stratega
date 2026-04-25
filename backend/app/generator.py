import os
import json
import re
import uuid
import pandas as pd
from dotenv import load_dotenv
from langchain_openai import ChatOpenAI
from langchain_core.prompts import ChatPromptTemplate
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from app.config import OUTPUTS_DIR

load_dotenv()

llm = ChatOpenAI(model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"))

Sarah_node = ChatPromptTemplate.from_template(
    """
    Your name is Sarah. You work as a content strategist at NeuraRank.
    Your role is to craft a weekly content calendar for NeuraRank.

    NeuraRank company details: {company_details}
    Weekly focus: {weekly_focus}

    Format the output for LinkedIn and Instagram.
    Create 6 rows total:
    - 3 days for LinkedIn
    - 3 days for Instagram
    - alternate platforms by day

    Return a table with 3 columns:
    Day, Topic, Platform
    """
)

john_node = ChatPromptTemplate.from_template(
    """
    Your name is John. You work as an associate content strategist at NeuraRank.

    Your role is to take Sarah's table and add a new column called Description.
    The Description column should contain a detailed guide for the content creator on how to properly achieve each deliverable.

    Return the updated table with:
    Day, Topic, Platform, Description

    Sarah's table:
    {Sarah_deliverable}
    """
)

sam_node = ChatPromptTemplate.from_template(
    """
    Your name is Sam. You work as a content editor at NeuraRank.

    Your role is to receive the deliverables from John and improve them for virality, while keeping them relevant to both our target audience and a wider audience.

    Details about NeuraRank:
    {NeuraRank}

    John's deliverables:
    {John_deliverable}
    """
)

mercy_node = ChatPromptTemplate.from_template(
    """
    You are a data cleaning assistant at NeuraRank.
    Convert the following messy table into valid JSON.

    Rules:
    - Return only valid JSON
    - Each row should be an object
    - Use consistent column names
    - Do not include explanations

    Messy table:
    {table}
    """
)

sarah_chain = Sarah_node | llm
john_chain = john_node | llm
sam_chain = sam_node | llm
mercy_chain = mercy_node | llm


def extract_last_json_block(text: str):
    if not text or not text.strip():
        raise ValueError("Empty output")

    blocks = re.findall(r"```json\s*(.*?)\s*```", text,
                        flags=re.DOTALL | re.IGNORECASE)
    if blocks:
        return json.loads(blocks[-1].strip())

    match = re.search(r"(\[.*\]|\{.*\})", text, flags=re.DOTALL)
    if not match:
        raise ValueError("No JSON found in output")

    return json.loads(match.group(1))


def sanitize_excel_filename(file_name: str | None):
    if not file_name or not file_name.strip():
        return f"content-calendar-{uuid.uuid4().hex[:8]}.xlsx"

    base_name = os.path.basename(file_name.strip())
    base_name = re.sub(r"[^\w\s.-]", "", base_name)
    base_name = re.sub(r"\s+", "-", base_name).strip(".-_")

    if not base_name:
        base_name = f"content-calendar-{uuid.uuid4().hex[:8]}"

    if not base_name.lower().endswith(".xlsx"):
        base_name = f"{base_name}.xlsx"

    return base_name


def get_unique_output_path(filename: str):
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    base, ext = os.path.splitext(filename)
    candidate = filename
    counter = 2

    while (OUTPUTS_DIR / candidate).exists():
        candidate = f"{base}-{counter}{ext}"
        counter += 1

    return candidate, str(OUTPUTS_DIR / candidate)


def normalize_calendar_dataframe(df: pd.DataFrame):
    column_map = {
        "day": "Day",
        "topic": "Topic",
        "platform": "Platform",
        "description": "Description",
    }
    renamed_columns = {
        column: column_map.get(str(column).strip().lower(), str(column).strip().title())
        for column in df.columns
    }
    normalized_df = df.rename(columns=renamed_columns)

    preferred_order = ["Day", "Topic", "Platform", "Description"]
    ordered_columns = [column for column in preferred_order if column in normalized_df.columns]
    ordered_columns += [column for column in normalized_df.columns if column not in ordered_columns]

    return normalized_df[ordered_columns].fillna("")


def save_formatted_excel(df: pd.DataFrame, requested_file_name: str | None):
    df = normalize_calendar_dataframe(df)
    filename = sanitize_excel_filename(requested_file_name)
    filename, filepath = get_unique_output_path(filename)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Content Calendar"
    worksheet.sheet_view.showGridLines = False

    headers = list(df.columns)
    worksheet.append(headers)

    for record in df.to_dict(orient="records"):
        worksheet.append([record.get(header, "") for header in headers])

    header_fill = PatternFill(fill_type="solid", fgColor="0F766E")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2EC"),
        right=Side(style="thin", color="D9E2EC"),
        top=Side(style="thin", color="D9E2EC"),
        bottom=Side(style="thin", color="D9E2EC"),
    )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 30

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_alignment
            cell.border = thin_border

    if worksheet.max_row > 1 and worksheet.max_column > 0:
        table_ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"
        table = Table(displayName="ContentCalendarTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    width_caps = {
        "Day": (14, 18),
        "Topic": (28, 42),
        "Platform": (16, 20),
        "Description": (45, 75),
    }

    for index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(index)
        values = [str(worksheet.cell(row=row, column=index).value or "") for row in range(1, worksheet.max_row + 1)]
        minimum, maximum = width_caps.get(header, (18, 45))
        calculated_width = max(minimum, min(maximum, max(len(value) for value in values) + 4))
        worksheet.column_dimensions[column_letter].width = calculated_width

    for row_index in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_index].height = 88

    workbook.save(filepath)

    return filename, filepath


def generate_content_calendar(company_details: str, weekly_focus: str, file_name: str | None = None):
    print("STEP 1: Starting Sarah chain...")
    sarah_outcome = sarah_chain.invoke({
        "company_details": company_details,
        "weekly_focus": weekly_focus
    }).content
    print("STEP 1 DONE")

    print("STEP 2: Starting John chain...")
    john_outcome = john_chain.invoke({
        "Sarah_deliverable": sarah_outcome
    }).content
    print("STEP 2 DONE")

    print("STEP 3: Starting Sam chain...")
    sam_outcome = sam_chain.invoke({
        "NeuraRank": company_details,
        "John_deliverable": john_outcome
    }).content
    print("STEP 3 DONE")

    print("STEP 4: Starting Mercy chain...")
    mercy_outcome = mercy_chain.invoke({
        "table": sam_outcome
    }).content
    print("STEP 4 DONE")

    print("STEP 5: Extracting JSON...")
    data = extract_last_json_block(mercy_outcome)
    print("STEP 5 DONE")

    print("STEP 6: Building DataFrame...")
    df = normalize_calendar_dataframe(pd.DataFrame(data))
    print("STEP 6 DONE")

    print("STEP 7: Saving Excel...")
    filename, filepath = save_formatted_excel(df, file_name)
    print("STEP 7 DONE")

    return {
        "df": df,
        "file_path": filepath,
        "file_name": filename,
        "sarah_outcome": sarah_outcome,
        "john_outcome": john_outcome,
        "sam_outcome": sam_outcome,
        "mercy_outcome": mercy_outcome,
    }
